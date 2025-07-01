import string
import streamlit as st
import pandas as pd
import re
from io import BytesIO
import openpyxl
import copy
import fitz

EXCEL_COLUMNS = [
    "STT", "NGÀY CHỨNG TỪ", "SỐ CHỨNG TỪ", "NGÀY HÓA ĐƠN", "SỐ HÓA ĐƠN",
    "TÊN MẶT HÀNG", "SỐ LƯỢNG", "ĐƠN VỊ", "GIÁ", "TỶ GIÁ", "THÀNH TIỀN NGUYÊN TỆ",
    "THÀNHTIỀN(VND)", "NỢ", "CÓ", "HẠNG MỤC", "TÊN", "Mã số thuế người bán", "GHI CHÚ"
]


def get_col_index(column_names, target_col):
    """Return 1-based index of the column matching target_col or None if not found."""
    for idx, col_name in enumerate(column_names):
        if col_name == target_col:
            return idx + 1
    return None


def get_cell_to_write(ws, row, col):
    """
    For merged cells, return the top-left cell of the merged region covering (row, col);
    otherwise, return the cell itself.
    """
    cell_coord = ws.cell(row=row, column=col).coordinate
    for merged_range in ws.merged_cells.ranges:
        if cell_coord in merged_range:
            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            return top_left_cell
    return ws.cell(row=row, column=col)


import fitz  # PyMuPDF


def extract_pdf_text(file):
    """Extract non-empty text lines from all pages of a PDF file stream,
       sorted in natural reading order by line-center then x."""
    doc = fitz.open(stream=file, filetype="pdf")
    lines_with_pos = []

    for page_num in range(doc.page_count):
        page = doc.load_page(page_num)
        blocks = page.get_text("blocks")
        # blocks: list of (x0, y0, x1, y1, text, block_no, block_type)

        for x0, y0, x1, y1, text, *_ in blocks:
            center_y = (y0 + y1) / 2
            for raw_line in text.splitlines():
                line = raw_line.strip()
                if not line:
                    continue
                lines_with_pos.append((center_y, x0, line))

    doc.close()

    # sort by vertical center, then by x (left→right)
    lines_with_pos.sort(key=lambda item: (item[0], item[1]))

    # strip out only the text, replace non‑breaking spaces
    sorted_lines = [line.replace("\xa0", " ") for _, _, line in lines_with_pos]
    return sorted_lines

def find_and_extract(lines, start_pattern="(.*)", pattern="(.*)", ignorance=None, post_processing=None, name=None):

    if filename=="TINVANG3703035442-C25TTV9.pdf":
        for line in lines:
            print(line)

    if post_processing==None:
        post_processing={
            "-" : "",
            " " : "",
        }

    start_search = re.compile(start_pattern)
    search_pattern = re.compile(pattern)

    result = ""
    found = False
    for i, line in enumerate(lines):
        if start_search.search(line):
            for j, l in enumerate(lines[i:]):
                parts = l.split(":")
                if (j==0 or parts[0].strip()=="") and len(parts)>1:
                    txt = parts[1].strip()
                else:
                    txt = l.strip()
                match = search_pattern.search(txt)
                if match:
                    result = match.group(1)
                    for k in post_processing:
                        result = result.replace(k, post_processing[k])
                    if ignorance and result in ignorance:
                        result = ""
                        break
                    else:
                        found = True
                        break
            if found:
                break

    if result=="":
        for line in lines:
            match = search_pattern.search(line.strip())
            if match:
                result = match.group(1)
                if ignorance and result in ignorance:
                    result = ""
                    continue
                for k in post_processing:
                    result = result.replace(k, post_processing[k])
                if name is None:
                    st.warning(f"The values found for file {filename} could be wrong. Please check")
                else:
                    st.warning(f"The value found for {name} for file {filename} could be wrong. Please check")
                break

    return result

def extract_seller_name(lines):

    def process_seller_name(str_):

        s = str_.split(":")[1] if ":" in str_ else str_
        s = str_.split()

        new_str = ""

        for i, w in enumerate(s):
            if w in string.punctuation or w.isupper():
                s = s[i:]
                break

        for w in s:
            if w in string.punctuation or w.isupper():
                new_str += f" {w}"
            else:
                break

        new_str = new_str.strip(string.punctuation).replace(" MST", "").strip()

        return new_str

    seller_name = ""
    try:
        for i, line in enumerate(lines):
            biline = lines[i].strip() + " " + lines[i + 1].strip()
            if ("CÔNG" in line and "TY" in line) and "MAI KA" not in biline:
                seller_name = process_seller_name(biline)
                break
    except:
        seller_name = find_and_extract(
            lines,
            r"Tên công ty|Company",
            r"(.*)",
            post_processing={},
        )

    return seller_name


def extract_other_values(lines):

    tax_code = find_and_extract(lines,
                                r"Mã số thuế|MST|Tax code|Tax Code",
                                r'^((?:\d\s*){10}(?:-\s*(?:\d\s*)+)?)$',
                                ["3700769325"],
                                name = "Mã số thuế")

    number = find_and_extract(lines,
                              r"Số(?:(?:\s|\([^)]*\))*)?:",
                              r'^(\d+)$',
                              name = "Số hóa đơn",
                              )

    return tax_code, number


def extract_tables_from_pdf(file):
    """
    Attempt to extract tables from first page of PDF using fitz's find_tables.
    Returns extracted table data rows starting from header row identified by 'STT'.
    """
    doc = fitz.open(stream=file.read(), filetype="pdf")
    page = doc[0]
    tables = page.find_tables()

    if not tables or not tables.tables:
        doc.close()
        return None

    table = tables[0].extract()
    doc.close()

    # Find header row containing 'STT'
    start_idx = 0
    for i, row in enumerate(table):
        if row and row[0] and "STT" in str(row[0]):
            start_idx = i
            break

    header_row = table[start_idx]
    chosen_cols_idx = [i for i, cell in enumerate(header_row) if cell]

    extracted_rows = []
    for row in table[start_idx:]:
        if not (row and row[0]):
            break
        filtered_row = [row[i] for i in chosen_cols_idx]
        extracted_rows.append(filtered_row)

    return extracted_rows

def postprocess_rows(item_table):

    try:
        start_idx = 2 if len(item_table) > 2 and item_table[2][0].strip()[0]=="1" else 1
        item_table = item_table[start_idx:]

        end_idx = len(item_table)
        for i, row in enumerate(item_table):
            if not row or not row[0] or not row[0].strip()[0].isdigit():
                end_idx = i
                break

        item_table = item_table[:end_idx]

        new_item_table = [[] for _ in range(len(item_table[0]))]
        for row in item_table:
            n_extract_row= len(row[0].strip().split("\n"))
            for i, cell in enumerate(row):
                cells = cell.strip().split("\n")
                new_item_table[i].extend(cells)
        new_item_table = [
            [
            new_item_table[j][i] for j in range(len(new_item_table))
            ]
            for i in range(len(new_item_table[0]))
        ]

        return new_item_table
    except Exception as e:
        return []

def parse_int(value_str):
    """Safely parse integer from string after removing dots, commas, spaces; return None on error."""
    try:
        normalized = value_str.replace('.', '').replace(',', '').replace(' ', '')
        return int(normalized)
    except Exception:
        return ''


filename = None
def extract_invoice_data_from_pdf(pdf_file_stream):
    """
    Main extraction function - fetches seller info, invoice number, dates,
    tax rate and item table rows from the given PDF stream.
    Returns list of dicts representing the invoice items.
    """
    # Reset stream pointer before multiple reads
    pdf_file_stream.seek(0)

    global filename
    filename = pdf_file_stream.name

    lines = extract_pdf_text(pdf_file_stream)

    # Reset stream to extract tables separately
    pdf_file_stream.seek(0)
    item_table = extract_tables_from_pdf(pdf_file_stream)

    if not item_table or len(item_table) == 0:
        raise ValueError("Cannot find table in the PDF")

    # Determine starting index for item rows
    item_rows = postprocess_rows(item_table)

    pdf_col_names = item_table[0]
    pdf_col_names = list(
        map(
            lambda x: " ".join(x.split("\n")[:-1]).strip() if len(x.split("\n")) > 1 else x,
            pdf_col_names
        )
    )

    if len(pdf_col_names) > 6:

        cols_map = {}
        if filename not in mapping_dict.keys():
            raise ValueError(f"File {filename} requires col_mapping but not found")
        for i, key in enumerate(["stt", "name", "unit", "qty", "price", "amount"]):
            cols_map[key] = mapping_dict[filename][i]

    else:
        cols_map = {"stt": 0, "name": 1, "unit": 2, "qty": 3, "price": 4, "amount": 5}

    seller_name = extract_seller_name(lines)
    tax_code, invoice_number = extract_other_values(lines)

    # Extract invoice date from lines
    date_pattern = re.compile(r'(\d{2}\s*[-/]\s*\d{2}\s*[-/]\s*\d{4})')
    invoice_date = ""
    for line in lines:
        match = date_pattern.search(line)
        if match:
            invoice_date = match.group(1).replace(" ", "").replace("-", "/")
            break

    # Extract tax rate (e.g. '10%') from lines
    tax_rate_pattern = re.compile(r"(?<!\d)(\d{1,3})\s*%(?!\d)")
    tax_rate = None
    for line in lines:
        match = tax_rate_pattern.search(line)
        if match:
            tax_rate = float(match.group(1)) / 100
            break

    if tax_rate is None:
        st.warning(f"Cannot find tax rate for {pdf_file_stream.name}, set tax_rate to 0")
        tax_rate = 0

    extracted_data = []
    for row in item_rows:
        # use the mapped indices instead of fixed positions
        try:
            stt        = str(row[cols_map["stt"]]).strip()
            item_name  = str(row[cols_map["name"]]).strip()
            unit       = str(row[cols_map["unit"]]).strip().lower()
            quantity   = parse_int(str(row[cols_map["qty"]]))
            unit_price = parse_int(str(row[cols_map["price"]]))
            amount     = parse_int(str(row[cols_map["amount"]]))
        except (IndexError, KeyError):
            # if mapping is invalid or row too short, skip
            continue

        if not all(var != "" for var in [amount]):
            raise ValueError("Missing thanh_tien")

        exchange_rate = 1  # currently hard-coded

        extracted_data.append({
            "STT": stt,
            "TÊN": seller_name,
            "Mã số thuế người bán": tax_code,
            "TÊN MẶT HÀNG": item_name,
            "ĐƠN VỊ": unit,
            "SỐ LƯỢNG": quantity,
            "GIÁ": unit_price,
            "TỶ GIÁ": exchange_rate,
            "THÀNH TIỀN NGUYÊN TỆ": amount,
            "THÀNHTIỀN(VND)": amount * float(exchange_rate),
            "Ghi chú": "",
            "Ký hiệu mẫu hóa đơn": "",
            "SỐ HÓA ĐƠN": invoice_number,
            "NGÀY HÓA ĐƠN": invoice_date,
            "NGÀY CHỨNG TỪ": invoice_date,
            "Thuế GTGT": amount * tax_rate
        })

    if len(extracted_data) == 0:
        raise ValueError("Cannot find table in the PDF_")
    else:
        extracted_data.append({
            "STT": 0,
            "TÊN": seller_name,
            "Mã số thuế người bán": tax_code,
            "TÊN MẶT HÀNG": "THUẾ GTGT",
            "ĐƠN VỊ": '',
            "SỐ LƯỢNG": '',
            "GIÁ": '',
            "TỶ GIÁ": 1,
            "THÀNH TIỀN NGUYÊN TỆ": sum(v["Thuế GTGT"] for v in extracted_data),
            "THÀNHTIỀN(VND)": sum(v["Thuế GTGT"]*float(v["TỶ GIÁ"]) for v in extracted_data),
            "Ghi chú": "",
            "Ký hiệu mẫu hóa đơn": "",
            "SỐ HÓA ĐƠN": invoice_number,
            "NGÀY HÓA ĐƠN": invoice_date,
            "NGÀY CHỨNG TỪ": invoice_date,
            "Thuế GTGT": '',
        })
    return extracted_data


def update_excel_with_data(pdf_files, excel_template_stream):
    """
    Processes multiple PDF files to extract invoice data, then updates the
    provided Excel template with the extracted rows preserving styles.
    Returns a BytesIO stream containing the updated Excel workbook.
    """
    all_items = []
    for pdf_file in pdf_files:
        try:
            data = extract_invoice_data_from_pdf(pdf_file)
            all_items.extend(data)
        except Exception as e:
            st.error(f"Error processing {pdf_file.name}: {str(e)}")

    # Renumber STT sequentially
    for idx, item in enumerate(all_items, start=1):
        item["STT"] = idx

    # Show extracted data in Streamlit app for preview
    df_preview = pd.DataFrame(all_items)
    st.write("### Extracted Data from PDF Files")
    st.dataframe(df_preview)

    # Load Excel workbook
    wb = openpyxl.load_workbook(excel_template_stream)
    ws = wb.active

    # Find Excel insert start row by detecting first row with numeric cell value in any column
    insert_row = None
    for r_idx, row_cells in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(val is not None and str(val).strip().isdigit() for val in row_cells):
            insert_row = r_idx
            break
    if insert_row is None:
        raise ValueError("Cannot find starting row to insert data in Excel sheet.")

    # Extract column names from header rows before insert_row
    col_names = [""] * (ws.max_column + 1)
    for row in ws.iter_rows(min_row=1, max_row=insert_row, values_only=True):
        for i, val in enumerate(row):
            if val and str(val).strip().replace("\n", "") in EXCEL_COLUMNS:
                col_names[i] = str(val).strip().replace("\n", "")

    # Copy styles of existing first data row for later re-application
    styles = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=insert_row, column=col_idx)
        styles.append({
            "font": copy.copy(cell.font),
            "border": copy.copy(cell.border),
            "fill": copy.copy(cell.fill),
            "number_format": copy.copy(cell.number_format),
            "protection": copy.copy(cell.protection),
            "alignment": copy.copy(cell.alignment)
        })

    # Determine end row of existing data by checking STT column for integers or None
    stt_col_idx = get_col_index(col_names, "STT")
    if stt_col_idx is None:
        raise ValueError("Cannot find 'STT' column in Excel sheet.")

    end_row = insert_row
    max_row = ws.max_row
    while end_row <= max_row:
        val = ws.cell(row=end_row, column=stt_col_idx).value
        if val is not None and not isinstance(val, int):
            break
        end_row += 1
    num_rows_to_delete = end_row - insert_row

    # Delete existing data rows
    if num_rows_to_delete > 0:
        ws.delete_rows(insert_row, amount=num_rows_to_delete)

    # Insert new rows for extracted data
    ws.insert_rows(insert_row, amount=len(all_items))

    # Fill new rows with data and apply styles
    for i, item in enumerate(all_items):
        row_num = insert_row + i
        for col_zero_idx in range(ws.max_column):
            col_name = col_names[col_zero_idx]
            val = item.get(col_name, "")
            cell = get_cell_to_write(ws, row_num, col_zero_idx + 1)
            style = styles[col_zero_idx]
            cell.font = style["font"]
            cell.border = style["border"]
            cell.fill = style["fill"]
            cell.number_format = style["number_format"]
            cell.protection = style["protection"]
            cell.alignment = style["alignment"]
            cell.value = val

    # Write totals in summary row (below last data row)
    summary_row = insert_row + len(all_items)
    col_amount_idx = get_col_index(col_names, "THÀNH TIỀN NGUYÊN TỆ")
    col_amount_vnd_idx = get_col_index(col_names, "THÀNHTIỀN(VND)")
    if not col_amount_idx or not col_amount_vnd_idx:
        raise ValueError("Missing required columns to write totals.")

    total_amount = sum(item.get("THÀNH TIỀN NGUYÊN TỆ", 0) for item in all_items)
    total_amount_vnd = sum(item.get("THÀNHTIỀN(VND)", 0) for item in all_items)

    get_cell_to_write(ws, summary_row, col_amount_idx).value = total_amount
    get_cell_to_write(ws, summary_row, col_amount_vnd_idx).value = total_amount_vnd

    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream


mapping_dict = {}
def main():
    st.title("PDF Invoice Data Processor")
    st.write("Upload PDF invoice files and an Excel template to process invoice data.")

    mapping_file = st.file_uploader(
        "Upload column-mapping TXT",
        type="txt",
        help='Format per line: "filename.pdf": 1,2,3,4,5,6'
    )
    global mapping_dict
    if mapping_file:
        text = mapping_file.getvalue().decode("utf-8")
        for line in text.splitlines():
            if not line.strip():
                continue
            parts = line.split(":")
            fname = parts[0].strip()
            indices = [int(s.strip())-1 for s in parts[1].split(",")]
            mapping_dict[fname] = indices

    pdf_files = st.file_uploader("Upload PDF Files", type="pdf", accept_multiple_files=True)
    excel_template = st.file_uploader("Upload Excel Template File", type="xlsx", accept_multiple_files=False)

    if pdf_files and excel_template:
            with st.spinner("Processing PDF files..."):
                try:
                    output_excel = update_excel_with_data(pdf_files, excel_template)
                    st.success("Processing complete!")
                    st.download_button(
                        label="Download Result Excel",
                        data=output_excel,
                        file_name="result.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as error:
                    st.error(f"Error during processing: {error}")


if __name__ == "__main__":
    main()
